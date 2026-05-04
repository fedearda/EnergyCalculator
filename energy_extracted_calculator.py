import argparse
from asammdf import MDF
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.styles import Font, PatternFill

def plot_energy(df):
    fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(12, 8), sharex=True)
    # --- Top plot: Power ---
    ax1.plot(df["timestamp"], df["Instant power [W]"]/1000,
            color="green", linewidth=1.5)
    ax1.set_ylabel("Power [kW]")
    ax1.set_title("Power, energy and SOC vs Time")
    ax1.grid(True, linestyle="--", alpha=0.6)
    ax1.axhline(0, color='black', linewidth=0.7, linestyle="-.")
    
    # --- Middle plot: Energy ---
    ax2.plot(df["timestamp"], df["Accumulated energy [kWh]"],
            color="blue", linewidth=2)
    ax2.set_ylabel("Energy [kWh]")
    ax2.grid(True, linestyle="--", alpha=0.6)

    # --- Bottom plot: SOC ---
    ax3.plot(df["timestamp"], df["SOC [%]"],
            color="red", linewidth=2)
    ax3.set_xlabel("Time [s]")
    ax3.set_ylabel("SOC [%]")
    ax3.grid(True, linestyle="--", alpha=0.6)
    plt.tight_layout()
    plt.show()

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input", help="Input MF4 file")
    parser.add_argument("--search", help="Search for signal name - the other args will be ignored", default=None, nargs='+')
    parser.add_argument("--output", help="Output Excel file", default=None)
    parser.add_argument("--start", type=float, default=None, help="Start time in seconds to cut data")
    parser.add_argument("--stop", type=float, default=None, help="End time in seconds to cut data")

    args = parser.parse_args()

    mdf = MDF(args.input)
    if args.search is not None:
        found = False
        for key in args.search:
            for name in mdf.channels_db:
                if name.__contains__(key):
                    print(f"{name}")
                    found = True
        if not found: print("No signal found")
        return
                    
    
    signals_list = "BMC_Strom", "BMC_Spannung", "BMC_TechnischerSOC"
    
    missing_signals = [s for s in signals_list if s not in mdf]
    
    if missing_signals:
        raise ValueError(f"The following signals were not found in the trace: {' ,'.join(missing_signals)}")
        
    df = mdf.to_dataframe(channels=signals_list)
    df.reset_index(names="timestamp",inplace=True)
    
    if args.start is not None:
        df = df[df["timestamp"] >= args.start]
        
    if args.stop is not None:
        df = df[df["timestamp"] <= args.stop]
        
    df["Instant power [W]"] = -df["BMC_Strom"]*df["BMC_Spannung"]
    df.rename(columns={"BMC_TechnischerSOC": "SOC [%]"}, inplace=True)
    
    t = df["timestamp"].values
    p = df["Instant power [W]"].values
    CONVERSION_FACTOR = 1/3600/1000
    energy = 0
    energy_series = [0]
    
    for i in range(1, len(p)):
        dt = t[i] - t[i-1]
        energy = energy + (p[i] + p[i-1]) * 0.5 * dt * CONVERSION_FACTOR
        energy_series.append(energy)
    
    df["Accumulated energy [kWh]"] = energy_series
    
    print(f"Total energy = {energy:.3f} kWh")
    
    if args.output is not None:
        with pd.ExcelWriter(args.output, engine="openpyxl",mode='w') as writer:
            df.to_excel(writer, index=False)
            ws = writer.sheets["Sheet1"]
            
            col = df.shape[1]+2
            if energy < 1:
                label_e = "Total energy [Wh]"
                energy = energy * 1000
            else:
                label_e = "Total energy [Wh]"
            ws.cell(row=1, column=col, value=label_e).font = Font(bold=True)
            ws.cell(row=2, column=col, value=energy).fill = PatternFill(start_color="00FF00", patternType='solid')
        print("Export completed")
        
    plot_energy(df)

if __name__ == "__main__":
    main()